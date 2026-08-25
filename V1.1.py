import os
import random
import smtplib
import time
import tkinter as tk
from email.header import Header
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import xlrd
from openpyxl import load_workbook


ACCOUNTS = [
    {"email": os.getenv("EMAIL1_ADDRESS", "1275716173@qq.com"), "password": os.getenv("EMAIL1_PASSWORD"), "smtp": "smtp.qq.com", "port": 465},
    {"email": os.getenv("EMAIL2_ADDRESS", "2521048381@qq.com"), "password": os.getenv("EMAIL2_PASSWORD"), "smtp": "smtp.qq.com", "port": 465},
]


def create_server(account):
    if not account["password"]:
        raise RuntimeError("未设置邮箱授权码。请先设置 EMAIL1_PASSWORD 和 EMAIL2_PASSWORD 环境变量。")
    server = smtplib.SMTP_SSL(account["smtp"], account["port"], timeout=30)
    server.login(account["email"], account["password"])
    return server


def send_email(server, receiver_email, subject, body, file_path, sender_email):
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = receiver_email
    message["Subject"] = Header(subject, "utf-8")
    message.attach(MIMEText(body, "plain", "utf-8"))
    with open(file_path, "rb") as attachment:
        part = MIMEApplication(attachment.read(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.add_header("Content-Disposition", "attachment", filename=Header(Path(file_path).name, "utf-8").encode())
    message.attach(part)
    server.send_message(message)


def normalize_name(value):
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_recipients(excel_path):
    if Path(excel_path).suffix.lower() == ".xlsx":
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = [(row[0], row[1]) for row in sheet.iter_rows(values_only=True) if len(row) >= 2]
        workbook.close()
        return rows
    sheet = xlrd.open_workbook(excel_path).sheets()[0]
    if sheet.ncols < 2:
        raise ValueError("Excel 至少需要两列：姓名、邮箱地址。")
    return [(sheet.cell(i, 0).value, sheet.cell(i, 1).value) for i in range(sheet.nrows)]


def start_sending():
    excel_path = excel_entry.get().strip()
    file_dir = folder_entry.get().strip()
    if not excel_path or not file_dir:
        messagebox.showerror("错误", "请选择收件人 Excel 和附件文件夹。")
        return
    try:
        recipients = load_recipients(excel_path)
    except Exception as exc:
        messagebox.showerror("读取失败", f"无法读取 Excel：{exc}")
        return
    account_index = 0
    server = None
    success = 0
    failed_names = []
    total = len(recipients)
    progress["maximum"] = total
    start_button.config(state="disabled")
    try:
        server = create_server(ACCOUNTS[account_index])
        for row_index, (raw_name, raw_email) in enumerate(recipients):
            name = normalize_name(raw_name)
            receiver_email = str(raw_email or "").strip()
            file_path = Path(file_dir) / f"{name}.xlsx"
            try:
                if not file_path.exists():
                    raise FileNotFoundError(f"未找到附件：{file_path.name}")
                send_email(server, receiver_email, "飞行二大队一中队工资明细", f"{name}您好，这是您本月的工资明细，请查收附件。", file_path, ACCOUNTS[account_index]["email"])
                success += 1
                log_text.insert(tk.END, f"发送成功：{name}\n")
            except Exception as exc:
                failed_names.append(name)
                log_text.insert(tk.END, f"发送失败：{name}（{exc}）\n")

            progress["value"] = row_index + 1
            percent = (row_index + 1) / total * 100 if total else 100
            percent_label.config(text=f"进度：{percent:.2f}%")
            root.update()
            if (row_index + 1) % 18 == 0 and row_index + 1 < total:
                server.quit()
                account_index = (account_index + 1) % len(ACCOUNTS)
                server = create_server(ACCOUNTS[account_index])
                time.sleep(15)
            elif row_index + 1 < total:
                time.sleep(random.uniform(8, 15))
                if (row_index + 1) % 10 == 0:
                    time.sleep(15)
    except Exception as exc:
        messagebox.showerror("发送中止", str(exc))
    finally:
        if server:
            try:
                server.quit()
            except Exception:
                pass
        start_button.config(state="normal")

    report_path = Path.cwd() / "失败名单.txt"
    report_path.write_text("\n".join(failed_names) if failed_names else "无失败记录", encoding="utf-8")
    messagebox.showinfo("发送结果", f"发送完成\n\n成功：{success} 封\n失败：{len(failed_names)} 封\n\n失败名单：{report_path}")


root = tk.Tk()
root.title("邮件自动发送工具")
root.geometry("560x480")
percent_label = tk.Label(root, text="进度：0.00%")
percent_label.pack(pady=(12, 4))

tk.Label(root, text="收件人 Excel：").pack()
excel_entry = tk.Entry(root, width=66)
excel_entry.pack(padx=12)
tk.Button(root, text="选择 Excel", command=lambda: excel_entry.insert(0, filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xls *.xlsx"), ("所有文件", "*.*")]))).pack(pady=4)

tk.Label(root, text="附件文件夹：").pack()
folder_entry = tk.Entry(root, width=66)
folder_entry.pack(padx=12)
tk.Button(root, text="选择附件文件夹", command=lambda: folder_entry.insert(0, filedialog.askdirectory())).pack(pady=4)

start_button = tk.Button(root, text="开始发送", command=start_sending, bg="#2e8b57", fg="white")
start_button.pack(pady=10)
log_text = tk.Text(root, height=12)
log_text.pack(fill="both", expand=True, padx=12)
progress = ttk.Progressbar(root, orient="horizontal", length=500, mode="determinate")
progress.pack(pady=12)
root.mainloop()
